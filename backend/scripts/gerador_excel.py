import mysql.connector
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import sys
import os

# Pega o caminho onde o arquivo será salvo (O Node.js vai passar isso)
output_path = sys.argv[1] if len(sys.argv) > 1 else 'Relatorio_KiSabor.xlsx'

try:
    # 1. Conecta no seu banco de dados MySQL
    conexao = mysql.connector.connect(
        host='localhost',
        user='root',
        password='',  # Coloque a senha do seu banco aqui
        database='kisabor'
    )

    # 2. Extrai os dados reais da sua tabela de pedidos (Hoje)
    query_pedidos = """
                    SELECT id, forma_pagamento, total, data_fechamento
                    FROM pedidos
                    WHERE status = 'Pago' AND DATE (data_fechamento) = CURDATE() \
                    """
    df_pedidos = pd.read_sql(query_pedidos, conexao)

    # 3. Extrai as Pizzas Mais Vendidas
    query_produtos = """
                     SELECT prod.nome as 'Sabor', COUNT(isab.id) as 'Quantidade'
                     FROM item_sabores isab
                              JOIN produtos prod ON isab.id_produto = prod.id
                              JOIN itens_pedido ip ON isab.id_item_pedido = ip.id
                              JOIN pedidos p ON ip.id_pedido = p.id
                     WHERE p.status = 'Pago' AND DATE (p.data_fechamento) = CURDATE()
                     GROUP BY prod.nome \
                     ORDER BY Quantidade DESC \
                     """
    df_produtos = pd.read_sql(query_produtos, conexao)

    # 4. Escreve no Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_pedidos.to_excel(writer, sheet_name='Vendas de Hoje', index=False)
        df_produtos.to_excel(writer, sheet_name='Ranking de Sabores', index=False)

    # 5. Deixa a planilha bonita (Estilo Premium)
    wb = openpyxl.load_workbook(output_path)
    fundo_cabecalho = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    fonte_cabecalho = Font(color='FFFFFF', bold=True)
    borda = Border(bottom=Side(style='thin', color='E2E8F0'))

    for aba in wb.sheetnames:
        ws = wb[aba]
        for celula in ws[1]:
            celula.fill = fundo_cabecalho
            celula.font = fonte_cabecalho
            celula.alignment = Alignment(horizontal='center')

        # Ajusta a largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 5

    # 6. Salva e fecha!
    wb.save(output_path)
    conexao.close()
    print(f"Sucesso: Planilha salva em {output_path}")

except Exception as e:
    print(f"Erro no Python: {e}")