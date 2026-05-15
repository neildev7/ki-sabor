import streamlit as st
import pandas as pd
import mysql.connector
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Ki-Sabor BI Intelligence", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
    <style>
    .main { background-color: #0f172a; }
    .stMetric { background-color: #1e293b; padding: 15px; border-radius: 10px; border: 1px solid #334155; }
    </style>
    """, unsafe_allow_html=True)

# --- CONEXÃO COM O BANCO ---
def conectar_banco():
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='neilnicoly13', 
        database='kisabor'
    )

@st.cache_data(ttl=30)
def carregar_dados():
    conn = conectar_banco()
    
    # 1. Vendas e Pagamentos
    query_vendas = """
        SELECT p.id, p.total, p.forma_pagamento, p.data_fechamento 
        FROM pedidos p WHERE p.status = 'Pago'
    """
    df_vendas = pd.read_sql(query_vendas, conn)
    df_vendas['data_fechamento'] = pd.to_datetime(df_vendas['data_fechamento'])
    
    # 2. Sabores
    query_sabores = """
        SELECT prod.nome as sabor, count(isab.id) as quantidade, p.data_fechamento
        FROM item_sabores isab
        JOIN produtos prod ON isab.id_produto = prod.id
        JOIN itens_pedido ip ON isab.id_item_pedido = ip.id
        JOIN pedidos p ON ip.id_pedido = p.id
        WHERE p.status = 'Pago'
        GROUP BY prod.nome, p.data_fechamento
    """
    df_sabores = pd.read_sql(query_sabores, conn)
    
    # 3. Avaliações
    query_feedbacks = """
        SELECT a.nota_pizza, a.nota_servico, a.comentario, p.data_fechamento
        FROM avaliacoes a
        JOIN pedidos p ON a.id_pedido = p.id
    """
    df_feedback = pd.read_sql(query_feedbacks, conn)
    
    conn.close()
    return df_vendas, df_sabores, df_feedback

# --- INÍCIO DA APLICAÇÃO ---
try:
    df_vendas, df_sabores, df_feedback = carregar_dados()

    # --- FILTROS (SIDEBAR) ---
    st.sidebar.title("Filtros Estratégicos")
    
    if not df_vendas.empty:
        min_date = df_vendas['data_fechamento'].min().date()
        max_date = df_vendas['data_fechamento'].max().date()
    else:
        min_date = pd.to_datetime("today").date()
        max_date = pd.to_datetime("today").date()
    
    start_date = st.sidebar.date_input("Início", min_date)
    end_date = st.sidebar.date_input("Fim", max_date)

    # Aplicando os filtros
    df_v_filt = df_vendas[(df_vendas['data_fechamento'].dt.date >= start_date) & (df_vendas['data_fechamento'].dt.date <= end_date)]
    df_s_filt = df_sabores[(df_sabores['data_fechamento'].dt.date >= start_date) & (df_sabores['data_fechamento'].dt.date <= end_date)]
    df_f_filt = df_feedback[(df_feedback['data_fechamento'].dt.date >= start_date) & (df_feedback['data_fechamento'].dt.date <= end_date)]

    # --- TÍTULO ---
    st.title("📊 Ki-Sabor BI Dashboard")

    # --- ABAS ---
    aba_financeiro, aba_produtos, aba_satisfacao, aba_exportar = st.tabs([
        "💰 Financeiro", "🍕 Produtos", "⭐ Satisfação", "📥 Exportar Dados"
    ])

    # --- ABA FINANCEIRO ---
    with aba_financeiro:
        col_f1, col_f2, col_f3 = st.columns(3)
        col_f1.metric("Faturamento Bruto", f"R$ {df_v_filt['total'].sum():,.2f}")
        col_f2.metric("Total de Pedidos", len(df_v_filt))
        col_f3.metric("Ticket Médio", f"R$ {df_v_filt['total'].mean():,.2f}" if len(df_v_filt)>0 else "R$ 0,00")

        c1, c2 = st.columns([2, 1])
        with c1:
            vendas_temporal = df_v_filt.groupby(df_v_filt['data_fechamento'].dt.date)['total'].sum().reset_index()
            fig_evolucao = px.area(vendas_temporal, x='data_fechamento', y='total', title="Evolução de Faturamento Diário",
                                  color_discrete_sequence=['#10b981'], line_shape='spline')
            st.plotly_chart(fig_evolucao, use_container_width=True)
        
        with c2:
            fig_meios = px.pie(df_v_filt, names='forma_pagamento', values='total', title="Meios de Pagamento", hole=0.5)
            st.plotly_chart(fig_meios, use_container_width=True)

    # --- ABA PRODUTOS ---
    with aba_produtos:
        st.subheader("Performance de Sabores")
        c3, c4 = st.columns(2)
        
        with c3:
            fig_tree = px.treemap(df_s_filt, path=['sabor'], values='quantidade', title="Volume de Vendas por Sabor",
                                  color='quantidade', color_continuous_scale='RdYlGn')
            st.plotly_chart(fig_tree, use_container_width=True)
            
        with c4:
            rank = df_s_filt.groupby('sabor')['quantidade'].sum().sort_values(ascending=True).reset_index()
            fig_rank = px.bar(rank, x='quantidade', y='sabor', orientation='h', title="Ranking de Sabores",
                              color='quantidade', color_continuous_scale='Viridis')
            st.plotly_chart(fig_rank, use_container_width=True)

    # --- ABA SATISFAÇÃO ---
    with aba_satisfacao:
        st.subheader("Experiência do Cliente")
        col_s1, col_s2 = st.columns(2)
        
        avg_pizza = df_f_filt['nota_pizza'].mean()
        avg_servico = df_f_filt['nota_servico'].mean()
        
        with col_s1:
            st.write("### ⭐ Média da Pizza")
            fig_pizza = go.Figure(go.Indicator(
                mode = "gauge+number", value = avg_pizza if not pd.isna(avg_pizza) else 0,
                domain = {'x': [0, 1], 'y': [0, 1]},
                gauge = {'axis': {'range': [0, 5]}, 'bar': {'color': "#10b981"}}
            ))
            fig_pizza.update_layout(height=250)
            st.plotly_chart(fig_pizza, use_container_width=True)

        with col_s2:
            st.write("### 🤵 Média do Atendimento")
            fig_serv = go.Figure(go.Indicator(
                mode = "gauge+number", value = avg_servico if not pd.isna(avg_servico) else 0,
                domain = {'x': [0, 1], 'y': [0, 1]},
                gauge = {'axis': {'range': [0, 5]}, 'bar': {'color': "#3b82f6"}}
            ))
            fig_serv.update_layout(height=250)
            st.plotly_chart(fig_serv, use_container_width=True)

        st.write("### 💬 Comentários Recentes")
        if not df_f_filt.empty:
            st.table(df_f_filt[['comentario']].dropna().tail(10))
        else:
            st.info("Ainda não há avaliações para o período selecionado.")

    # --- ABA EXPORTAR DADOS ---
    with aba_exportar:
        st.subheader("🚀 Central de Relatórios Inteligentes")
        st.write("Gere um arquivo consolidado com todos os dados financeiros, performance de produtos e feedback dos clientes.")

        df_resumo = pd.DataFrame({
            "Indicador": ["Período Analisado", "Faturamento Bruto", "Total de Pedidos", "Ticket Médio", "Média Satisfação Pizza", "Média Satisfação Atendimento"],
            "Valor": [
                f"{start_date} até {end_date}",
                f"R$ {df_v_filt['total'].sum():,.2f}",
                len(df_v_filt),
                f"R$ {df_v_filt['total'].mean():,.2f}" if len(df_v_filt) > 0 else "R$ 0,00",
                f"{df_f_filt['nota_pizza'].mean():.2f}" if not df_f_filt.empty else "N/A",
                f"{df_f_filt['nota_servico'].mean():.2f}" if not df_f_filt.empty else "N/A"
            ]
        })

        def to_excel_premium(df_res, df_fin, df_prod, df_sat):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_res.to_excel(writer, sheet_name='Resumo Executivo', index=False)
                df_fin.to_excel(writer, sheet_name='Detalhamento Financeiro', index=False)
                df_prod.to_excel(writer, sheet_name='Ranking de Sabores', index=False)
                df_sat.to_excel(writer, sheet_name='Feedback dos Clientes', index=False)
                
                workbook = writer.book
                
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except: pass
                        worksheet.column_dimensions[column_letter].width = max_length + 2
            
            return output.getvalue()

        try:
            excel_data = to_excel_premium(df_resumo, df_v_filt, df_s_filt, df_f_filt)
            
            st.download_button(
                label="📥 Baixar Relatório Gerencial Premium (.xlsx)",
                data=excel_data,
                file_name=f"Relatorio_Executivo_KiSabor_{start_date}_a_{end_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Clique para baixar o arquivo Excel formatado com todas as métricas."
            )
            
            st.success("Relatório gerado com sucesso! Clique no botão acima para baixar.")
            
        except Exception as e:
            st.error(f"Erro ao formatar Excel: {e}")

except Exception as e:
    st.error(f"Erro ao carregar dados: {e}")
    st.info("Verifique se o MySQL está rodando e as tabelas estão corretas.")